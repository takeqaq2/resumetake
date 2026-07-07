import paramiko
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

SSH_HOST = '118.145.76.118'
SSH_PORT = 10000
SSH_USER = 'wgy'
SSH_PASS = 'a2154321..'
MYSQL_USER = 'readOnly'
MYSQL_PASS = '666888'

CITY_MAP = {
    "2": "北京-云桌面", "3": "广州-云桌面", "4": "重庆-云桌面",
    "5": "上海-云桌面", "6": "南京-云桌面", "7": "成都-云桌面",
    "8": "北京", "9": "呼和浩特", "10": "广州", "11": "苏州-停用",
    "12": "济南", "13": "杭州", "14": "苏州", "15": "芜湖",
    "16": "上海", "17": "杭州-停用", "33": "贵阳", "19": "重庆",
    "20": "合肥", "21": "南京", "22": "惠州", "23": "佛山",
    "25": "武汉", "26": "兰州", "27": "成都", "28": "西宁",
    "29": "昆明", "30": "郑州", "31": "西安", "32": "长沙"
}

WORK_DIR = 'F:\\project2take\\testmysql'
TODAY = datetime.now().strftime('%Y-%m-%d')


def get_ssh():
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(SSH_HOST, port=SSH_PORT, username=SSH_USER, password=SSH_PASS, timeout=15)
    return ssh


def query(ssh, sql):
    cmd = f'mysql -h 127.0.0.1 -u {MYSQL_USER} -p{MYSQL_PASS} app -e "{sql}"'
    stdin, stdout, stderr = ssh.exec_command(cmd, timeout=30)
    return stdout.read().decode()


def query_dicts(ssh, sql):
    lines = query(ssh, sql).strip().split('\n')
    if len(lines) < 2:
        return []
    headers = lines[0].split('\t')
    rows = []
    for line in lines[1:]:
        vals = line.split('\t')
        row = {}
        for i, h in enumerate(headers):
            row[h] = vals[i] if i < len(vals) else ''
        rows.append(row)
    return rows


def main():
    ssh = get_ssh()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    print(f'=== 冷却期IP更换查询 {now} ===\n')

    # 1. canceling -> canceled
    print('--- 1. canceling 状态实例 ---')
    canceling = query_dicts(ssh, """
        SELECT pi.id, pi.node_id, pi.ip_id, pi.status, pi.canceled_at, pi.created_at,
            i.public_address as ip_addr, n.address as node_addr, l.city, l.id as line_id
        FROM proxy_instance pi
        LEFT JOIN ip i ON pi.ip_id = i.id
        LEFT JOIN node n ON pi.node_id = n.id
        LEFT JOIN line l ON n.line_id = l.id
        WHERE pi.status = 'canceling' AND pi.deleted_at = '0001-01-01 00:00:00'
    """)
    print(f'  找到 {len(canceling)} 条')

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    sql_lines = []
    for r in canceling:
        sql_lines.append(f"UPDATE proxy_instance SET status='canceled', updated_at='{now_str}' WHERE id={r['id']};")

    # 2. 查询冷却期IP（存活>72h且2个月内退订的）
    print('\n--- 2. 冷却期IP查询 ---')
    cooldown = query_dicts(ssh, f"""
        SELECT
            ip.id as ip_id,
            ip.line_id,
            l.city,
            ip.node_id,
            CONCAT('一拖', node_ip_stat.node_ip_count) as 一拖几,
            node_ip_stat.node_ip_count as 同node有效IP数,
            n.address as node_address,
            ip.public_address,
            ip.iface_address,
            ip.mac_address,
            '冷却期内' as 冷却状态,
            cooling_stat.cooling_canceled_at as 冷却开始时间,
            cooling_stat.cooldown_until as 冷却结束时间,
            ip.created_at as ip_created_at
        FROM ip
        LEFT JOIN (
            SELECT node_id, COUNT(*) as node_ip_count
            FROM ip WHERE deleted_at = '0001-01-01 00:00:00'
            GROUP BY node_id
        ) node_ip_stat ON node_ip_stat.node_id = ip.node_id
        LEFT JOIN (
            SELECT
                pi1.ip_id,
                MAX(pi1.canceled_at) as cooling_canceled_at,
                DATE_ADD(MAX(pi1.canceled_at), INTERVAL 2 MONTH) as cooldown_until
            FROM proxy_instance pi1
            WHERE pi1.deleted_at = '0001-01-01 00:00:00'
              AND pi1.canceled_at != '0001-01-01 00:00:00'
              AND pi1.canceled_at > DATE_ADD(pi1.created_at, INTERVAL 72 HOUR)
              AND DATE_ADD(pi1.canceled_at, INTERVAL 2 MONTH) > NOW()
            GROUP BY pi1.ip_id
        ) cooling_stat ON cooling_stat.ip_id = ip.id
        LEFT JOIN node n ON n.id = ip.node_id AND n.deleted_at = '0001-01-01 00:00:00'
        LEFT JOIN line l ON l.id = ip.line_id AND l.deleted_at = '0001-01-01 00:00:00'
        WHERE ip.deleted_at = '0001-01-01 00:00:00'
          AND cooling_stat.ip_id IS NOT NULL
        ORDER BY ip.line_id, ip.node_id, ip.id
    """)
    print(f'  找到 {len(cooldown)} 个冷却期IP')

    # 3. 按城市统计
    print('\n--- 3. 按城市统计冷却期IP ---')
    city_stats = {}
    for r in cooldown:
        city = r.get('city', '未知')
        if city not in city_stats:
            city_stats[city] = []
        city_stats[city].append(r)

    for city, items in sorted(city_stats.items()):
        print(f'  {city}: {len(items)} 个冷却期IP')

    # 4. 生成 xlsx
    print('\n--- 4. 生成 xlsx ---')
    wb = Workbook()

    # Sheet1: 冷却期IP列表
    ws1 = wb.active
    ws1.title = '冷却期IP-需更换'
    headers1 = ['城市', 'line_id', 'node_id', '一拖几', '同node有效IP数',
                 '节点地址', 'IP地址', 'iface地址', 'MAC地址',
                 '冷却开始时间', '冷却结束时间', 'IP入库时间']
    ws1.append(headers1)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for r in cooldown:
        ws1.append([
            r.get('city', ''), r.get('line_id', ''), r.get('node_id', ''),
            r.get('一拖几', ''), r.get('同node有效IP数', ''),
            r.get('node_address', ''), r.get('public_address', ''),
            r.get('iface_address', ''), r.get('mac_address', ''),
            r.get('冷却开始时间', ''), r.get('冷却结束时间', ''),
            r.get('ip_created_at', '')
        ])

    # 高亮重复node_id（同一node的IP标黄）
    node_seen = {}
    for row_idx in range(2, ws1.max_row + 1):
        nid = str(ws1.cell(row=row_idx, column=3).value)
        if nid not in node_seen:
            node_seen[nid] = 0
        node_seen[nid] += 1

    for row_idx in range(2, ws1.max_row + 1):
        nid = str(ws1.cell(row=row_idx, column=3).value)
        if node_seen.get(nid, 0) > 1:
            for col in range(1, len(headers1) + 1):
                ws1.cell(row=row_idx, column=col).fill = yellow_fill

    # Sheet2: 按城市汇总
    ws2 = wb.create_sheet('按城市汇总')
    ws2.append(['城市', '冷却期IP数量'])
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=1).font = header_font
    ws2.cell(row=1, column=2).font = header_font
    for city, items in sorted(city_stats.items()):
        ws2.append([city, len(items)])
    ws2.append(['总计', len(cooldown)])

    # Sheet3: SQL
    ws3 = wb.create_sheet('执行SQL')
    ws3.append(['说明', 'SQL'])
    ws3.cell(row=1, column=1).fill = header_fill
    ws3.cell(row=1, column=2).fill = header_fill
    ws3.cell(row=1, column=1).font = header_font
    ws3.cell(row=1, column=2).font = header_font
    ws3.append(['canceling->canceled', f'共 {len(sql_lines)} 条UPDATE'])
    for i, s in enumerate(sql_lines, 1):
        ws3.append([f'{i}', s])

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 100

    for ws in [ws1]:
        for col in range(1, ws.max_column + 1):
            max_len = max(len(str(ws.cell(row=r, column=col).value or '')) for r in range(1, ws.max_row + 1))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 30)

    output_path = f'{WORK_DIR}\\冷却期IP更换_{TODAY}.xlsx'
    wb.save(output_path)
    print(f'  已保存: {output_path}')

    # 5. 输出SQL
    print('\n--- 5. canceling->canceled SQL ---')
    for s in sql_lines:
        print(s)

    ssh.close()
    print('\n=== 完成 ===')


if __name__ == '__main__':
    main()
